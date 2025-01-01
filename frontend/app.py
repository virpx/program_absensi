import time
import pyqrcode
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import font as tkFont
from tkinter import font
from tkinter import messagebox ,filedialog
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import io
import re
import pandas
import requests
from tkinter import filedialog
from openpyxl import load_workbook
import json
import base64
import pandas as pd

# Inisialisasi driver di luar fungsi agar dapat diakses secara global
driver = None
# pilih_aksi = tk.StringVar()
btnsave = None
list_entry_add_siswa = []
tokenlogin = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3MzU3MDAxNzIsImRhdGEiOiJhZG1pbmxvZ2luIiwiaWF0IjoxNzM1NjEzNzcyfQ.iTkBpsCsRi_lvOr-NpvPRPH67U9kj0_zMmBaSsqn4HI"
# Fungsi untuk mengecek login setiap beberapa detik
def check_login_status():
    if is_logged_in():
        qr_frame.pack_forget()
        absen_frame.pack(fill="both", expand=True)
    else:
        # Cek lagi dalam 5 detik
        root.after(5000, check_login_status)

# Fungsi yang dipanggil saat jendela Tkinter ditutup
def on_closing():
    global driver
    if driver:
        driver.quit()
    root.destroy() 

# Cek Login WA
def is_logged_in():
    try:
        # Coba cari elemen yang hanya ada jika login berhasil
        chat_list = driver.find_elements("css selector", "#side")
        return len(chat_list) > 0
    except Exception as e:
        return False

# Array berisi data siswa
data_siswa = [
    {"nisn": "1234567890", "nama": "Ahmad Fauzi", "no_hp_ortu": "082140950288"},
    {"nisn": "9876543210", "nama": "Siti Nurhaliza", "no_hp_ortu": "082140950288"},
    {"nisn": "1122334455", "nama": "Budi Santoso", "no_hp_ortu": "082140950288"},
    {"nisn": "5566778899", "nama": "Rina Agustina", "no_hp_ortu": "082140950288"},
    {"nisn": "6677889900", "nama": "Yusuf Kurniawan", "no_hp_ortu": "082140950288"},
]

# Cek qrcode absen

def on_key_release(event):
    current_text = nik_entry.get()

    if len(current_text) == 10:  # Cek jika panjang input 10 karakter
        siswa_ditemukan = next((siswa for siswa in data_siswa if siswa["nisn"] == current_text), None)

        if siswa_ditemukan:
            nama = siswa_ditemukan["nama"]
            no_hp = siswa_ditemukan["no_hp_ortu"]

            try:
                driver.get(f"https://web.whatsapp.com/send?phone=62{no_hp}&text=Selamat pagi Bapak/Ibu, kami ingin menyampaikan bahwa siswa {nama} telah hadir di sekolah.")

                # Tunggu beberapa saat agar pengguna dapat login
                time.sleep(15)

                # Klik tombol kirim pada WhatsApp Web (dengan Xpath)
                try:
                    send_button = driver.find_element("css selector", 'button[aria-label="Kirim"]')
                    send_button.click()
                    nama = ""
                    no_hp = ""
                except Exception as e:
                    messagebox.showerror("Error", f"Gagal menekan tombol kirim: {str(e)}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Gagal membuka WhatsApp Web: {str(e)}")
        else:
            messagebox.showwarning("NISN Tidak Ditemukan", "NISN tidak ditemukan!")
        
        nik_entry.delete(0, tk.END)

def sanitize_phone_number(phone):
    phone = re.sub(r'\D', '', phone)
    if phone.startswith('+62'):
        phone = phone[3:]  
    elif phone.startswith('62'):
        phone = phone[2:]  
    elif phone.startswith('0'):
        phone = phone[1:]  
    return str(phone)



# Fungsi untuk membuat tombol melingkar
def create_rounded_button(canvas, x, y, width, height, radius, text, command=None):
    # Gambar bentuk bulat di sekitar tombol
    canvas.create_oval(x, y, x + radius*2, y + radius*2, fill="purple", outline="")
    canvas.create_oval(x + width - radius*2, y, x + width, y + radius*2, fill="purple", outline="")
    canvas.create_oval(x, y + height - radius*2, x + radius*2, y + height, fill="purple", outline="")
    canvas.create_oval(x + width - radius*2, y + height - radius*2, x + width, y + height, fill="purple", outline="")
    
    # Menggambar persegi panjang di antara lingkaran untuk menyelesaikan tampilan melingkar
    canvas.create_rectangle(x + radius, y, x + width - radius, y + height, fill="purple", outline="")
    canvas.create_rectangle(x, y + radius, x + width, y + height - radius, fill="purple", outline="")

    # Menambahkan teks pada tombol
    button_text = canvas.create_text(x + width / 2, y + height / 2, text=text, fill="white", font=("Arial", 12, "bold"))

    # Menambahkan event klik pada tombol
    if command:
        canvas.tag_bind(button_text, "<Button-1>", lambda event: command())
        canvas.tag_bind("button_background", "<Button-1>", lambda event: command())

# Fungsi untuk memuat gambar dari folder assets
def load_image(path, width, height):
    image = Image.open(path)
    image = image.resize((width, height), Image.LANCZOS)
    return ImageTk.PhotoImage(image)

# Membuat antarmuka GUI menggunakan Tkinter
root = tk.Tk()
root.title("WhatsApp Web QR Code Fetcher")

# Ukuran window
root.geometry("1000x1000")

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame Login # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


def login():
    userid = email_entry.get()
    password = password_entry.get()
    payload = {
        "username": userid,
        "password":password
    }
    response = requests.post("http://localhost:3000/login", data=payload)
    datalogin = json.loads(response.text)
    if datalogin["success"] == 0:
        messagebox.showerror("Login Failed", datalogin["data"])
    else:
        # Pindah ke frame QR Code
        tokenlogin = datalogin["data"]
        email_entry.delete(0, tk.END)
        password_entry.delete(0, tk.END)
        login_frame.pack_forget()
        menu_frame.pack(fill="both", expand=True)
        # fetch_and_show_qr_code()

login_frame = tk.Frame(root, bg="white")
# login_frame.pack(fill="both", expand=True)

# Sub-frame untuk memusatkan isi
center_frame = tk.Frame(login_frame, bg="white")
center_frame.pack(expand=True, padx=20, pady=20)

# Gambar di tengah
try:
    original_image = Image.open("assets/login_image.png")
    resized_image = original_image.resize((250, 250))  # Sesuaikan ukuran gambar
    photo = ImageTk.PhotoImage(resized_image)
    image_label = tk.Label(center_frame, image=photo, borderwidth=0, highlightthickness=0, bg="white")
    image_label.grid(row=0, column=0, padx=20, rowspan=3)  # Gambar menempati posisi kiri
except Exception as e:
    error_label = tk.Label(center_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.grid(row=0, column=0, padx=20, rowspan=3)

# Sub-frame kanan untuk form login
form_frame = tk.Frame(center_frame, bg="white")
form_frame.grid(row=0, column=1, padx=20, sticky="n")

# Judul
label_font = tkFont.Font(family="Helvetica", size=20, weight="bold")
title_label = tk.Label(form_frame, text="Login as Admin", font=label_font, bg="white", fg="#6F42C1")
title_label.pack(pady=(20, 20))

# Frame untuk input email
email_frame = tk.Frame(form_frame, bg="white")
email_frame.pack(pady=(10, 10), fill="x")

# Email icon
email_icon = tk.Label(email_frame, text="👤", font=("Helvetica", 14), bg="white", fg="#6F42C1")
email_icon.pack(side=tk.LEFT, padx=(10, 10))

# Email entry
email_entry = tk.Entry(
    email_frame,
    font=("Helvetica", 14),
    bg="#f0f0ff",
    fg="black",
    insertbackground="black",
    highlightthickness=2,
    highlightbackground="#a0a0ff",
    relief="solid",
    bd=1
)
email_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 10), ipady=5)

# Frame untuk input password
password_frame = tk.Frame(form_frame, bg="white")
password_frame.pack(pady=(10, 20), fill="x")

# Password icon
password_icon = tk.Label(password_frame, text="🔒", font=("Helvetica", 14), bg="white", fg="#6F42C1")
password_icon.pack(side=tk.LEFT, padx=(10, 10))

# Password entry
password_entry = tk.Entry(
    password_frame,
    font=("Helvetica", 14),
    bg="#f0f0ff",
    fg="black",
    insertbackground="black",
    highlightthickness=2,
    highlightbackground="#a0a0ff",
    relief="solid",
    bd=1,
    show="*"
)
password_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 10), ipady=5)

# Login button
button_frame = tk.Frame(form_frame, bg="white")  # Sub-frame untuk memisahkan tombol
button_frame.pack(pady=(5, 0), fill="x")  # Tambahkan padding di atas tombol

# Tombol login
login_btn = tk.Canvas(button_frame, width=110, height=50, bg="white", highlightthickness=0)
login_btn.pack(anchor="center", pady=(4, 0))  # Centering tombol login dan beri padding bawah
create_rounded_button(
    login_btn,
    x=5,
    y=5,
    width=100,
    height=40,
    radius=20,
    text="LOGIN",
    command=login
)

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame QR Code # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# Fungsi untuk kembali ke halaman login
def go_back():
    qr_frame.pack_forget()
    login_frame.pack(fill="both", expand=True)

# Fungsi untuk mengambil dan menampilkan QR code di Tkinter
def fetch_and_show_qr_code():
    global driver, qr_label
    try:
        # Konfigurasi headless browser menggunakan Chrome
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument('--disable-dev-shm-usage')

        # Inisialisasi WebDriver
        driver = webdriver.Chrome(options=chrome_options)
        driver.get("https://web.whatsapp.com/")
        
        # Tunggu sebentar agar halaman selesai dimuat
        time.sleep(4)
        
        # Tunggu elemen dengan atribut data-ref muncul
        qr_element = driver.find_element("css selector", "div[data-ref]")

        # Dapatkan nilai dari atribut data-ref
        data_ref = qr_element.get_attribute("data-ref")
        
        # Buat QR code dari data-ref
        qr_code = pyqrcode.create(data_ref)

        # Simpan QR code ke dalam stream gambar PNG
        buffer = io.BytesIO()
        qr_code.png(buffer, scale=6)
        buffer.seek(0)

        # Buka gambar menggunakan PIL dan konversi ke format yang bisa digunakan di Tkinter
        qr_image = Image.open(buffer)
        qr_image_tk = ImageTk.PhotoImage(qr_image)

        # Jika ada QR code yang sudah ditampilkan sebelumnya, hapus
        if qr_label is not None:
            qr_label.destroy()

        # Label untuk menampilkan gambar QR code di Tkinter
        qr_label = tk.Label(qr_frame, image=qr_image_tk)
        qr_label.image = qr_image_tk  
        qr_label.pack(pady=20)

    except Exception as e:
        messagebox.showerror("Error", f"Gagal mengambil QR Code: {str(e)}")


qr_frame = tk.Frame(root , bg="white")

# Variabel untuk label gambar QR code
qr_label = None

# Menambahkan deskripsi
description_label = tk.Label(qr_frame, text="Scan the QR code above to proceed.", 
                              bg="white", fg="black", font=("Arial", 10))
description_label.pack(pady=10)

# Tombol Back di halaman QR Code
back_button = tk.Button(qr_frame, text="Back", command=go_back , bg="#4CAF50", fg="black", font=("Arial", 12, "bold"))
back_button.pack(pady=20)

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame Menu # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def todata():
    menu_frame.pack_forget()
    databases_frame.pack(fill="both", expand=True)

# Fungsi untuk ke halaman report
def to_report():
    menu_frame.pack_forget()
    report_frame.pack(fill="both", expand=True)

# Fungsi untuk exit       
def exite():
    menu_frame.pack_forget()
    login_frame.pack(fill="both", expand=True)

# Fungsi untuk menampilkan QR Code dan cek login       
def qrcode():
    menu_frame.pack_forget()
    qr_frame.pack(fill="both", expand=True)
    fetch_and_show_qr_code()
    check_login_status()

menu_frame = tk.Frame(root, bg="white")
# menu_frame.pack(fill="both", expand=True)

# Palet warna utama
primary_color = "#6A1B9A"  # Ungu untuk aksen utama
background_color = "white"

# Section Tengah (terdiri dari kiri, tengah, dan kanan)
tengah_frame = tk.Frame(menu_frame, bg="white")
tengah_frame.place(relx=0.5, rely=0.4, anchor="center", relwidth=0.9, relheight=0.5)

# Bagian Kiri di Tengah
kiri_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
kiri_frame.pack(fill="y", side="left", padx=10, pady=10)

# Bagian Tengah di Tengah
tengah_tengah_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
tengah_tengah_frame.pack(fill="both", side="left", expand=True, padx=10, pady=10)

# Bagian Kanan di Tengah
kanan_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
kanan_frame.pack(fill="y", side="right", padx=10, pady=10)

# Section Bawah
bawah_frame = tk.Frame(menu_frame, bg=background_color, height=80)
bawah_frame.pack(fill="x", side="bottom")

# Judul Menu
title_label = tk.Label(menu_frame, text="Main Menu", bg=background_color, fg="Black", font=("Arial", 24, "bold"))
title_label.pack(pady=10)

# Load Gambar
try:
    image_path = "assets/database.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(kiri_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(kiri_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Database di halaman Menu
database = tk.Canvas(kiri_frame, width=110, height=50, bg="white", highlightthickness=0)
database.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(database, x=5, y=5, width=100, height=40, radius=20, text="Database", command=todata)

# Load Gambar
try:
    image_path = "assets/reicon.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(tengah_tengah_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(tengah_tengah_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Riport di halaman Menu
riport = tk.Canvas(tengah_tengah_frame, width=110, height=50, bg="white", highlightthickness=0)
riport.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(riport, x=5, y=5, width=100, height=40, radius=20, text="Riport", command=to_report)

# Load Gambar
try:
    image_path = "assets/abs.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(kanan_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(kanan_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Absensi Siswa di halaman Menu
absensi = tk.Canvas(kanan_frame, width=110, height=50, bg="white", highlightthickness=0)
absensi.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(absensi, x=5, y=5, width=100, height=40, radius=20, text="Absensi Siswa", command=qrcode)

# Tombol Exit di pojok kanan bawah pada bawah_frame
exited_canvas = tk.Canvas(bawah_frame, width=100, height=50, bg="white", highlightthickness=0)
exited_canvas.place(relx=0.85, rely=0.25, anchor="center") 
create_rounded_button(exited_canvas, x=5, y=5, width=90, height=40, radius=20, text="Exit", command=exite)

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame Absensi # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# Fungsi untuk kembali ke halaman login
def go_login():
    absen_frame.pack_forget()
    login_frame.pack(fill="both", expand=True)    
def on_key_release(e):
    global tokenlogin
    if len(nik_entry.get()) == 10:
        payload = {
            "login": tokenlogin,
            "data":nik_entry.get()
        }
        response = requests.post("http://localhost:3000/absen", data=payload)
        datalogin = json.loads(response.text)
        if datalogin['success'] == 0:
            messagebox.showerror("Gagal Absen", datalogin["data"])
        nik_entry.delete(0,tk.END)
absen_frame = tk.Frame(root,bg="white")
# absen_frame.pack(fill="both", expand=True)

# Membuat dua section di dalam frame absensi

# Membuat frame atas dan bawah untuk menambah margin vertikal pada garis pemisah
top_margin = tk.Frame(absen_frame, height=20, bg="white")  # margin atas
top_margin.pack(fill="x")
section_kiri = tk.Frame(absen_frame, width=550, bg="white")
section_kiri.pack(side="left", fill="both", expand=True)

section_kanan = tk.Frame(absen_frame, width=250, bg="white")
section_kanan.pack(side="right", fill="both", expand=True)

# Membuat garis tegak lurus berwarna ungu sebagai pemisah
separator = tk.Canvas(absen_frame, width=20, height=560, bg="white", highlightthickness=0)
separator.create_line(10, 0, 10, 560, fill="purple", width=5)  # Garis ungu di tengah canvas
separator.pack(side="left", padx=10, pady=20)

# Isi section kiri

# NIK scan
nik = tk.Label(section_kiri, text="Barcode Scan", font=("Arial", 26, "bold"), fg="black", bg="white" )
nik.pack(pady=(100, 10), padx=(10, 0),anchor="center")

# Load Gambar
try:
    image_path = "assets/scan.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 350, 350)  # Mengatur ukuran gambar
    img_label = tk.Label(section_kiri, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(section_kiri, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Isi section kanan

# Entry untuk input nomor
nik_entry = tk.Entry(
    section_kanan,
    font=("Helvetica", 14),       
    bg="#f0f0ff",                  
    fg="black",                    
    insertbackground="black",      
    highlightthickness=2,          
    highlightbackground="#a0a0ff", 
    relief="solid",                
    bd=1                           
)

nik_entry.place(relx=0.5, rely=0.4, anchor="center", width=200, height=30)  # Ukuran dan posisi

# Bind event key release pada entry
nik_entry.bind("<KeyRelease>", on_key_release)

# Membuat Canvas untuk tombol di section kanan
back_canvas = tk.Canvas(section_kanan, width=100, height=50, bg="white", highlightthickness=0)
back_canvas.place(relx=0.5, rely=0.6, anchor="center")  # Posisi di tengah secara horizontal, sedikit di bawah tengah secara vertikal

# Membuat tombol "Back" melingkar
create_rounded_button(back_canvas, x=5, y=5, width=90, height=40, radius=20, text="Back", command=go_login)

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame Database # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

databases_frame = tk.Frame(root)

# Palet warna utama
primary_color = "#6A1B9A"  # Ungu untuk aksen utama
background_color = "white"

# Section Tengah (terdiri dari kiri, tengah, dan kanan)
tengah_frame = tk.Frame(databases_frame, bg="white")
tengah_frame.place(relx=0.5, rely=0.4, anchor="center", relwidth=0.9, relheight=0.5)

# Bagian Kiri di Tengah
kiri_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
kiri_frame.pack(fill="y", side="left", padx=10, pady=10)

# Bagian Tengah di Tengah
tengah_tengah_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
tengah_tengah_frame.pack(fill="both", side="left", expand=True, padx=10, pady=10)

# Bagian Kanan di Tengah
kanan_frame = tk.Frame(tengah_frame, bg=background_color, width=200)
kanan_frame.pack(fill="y", side="right", padx=10, pady=10)

# Section Bawah
bawah_frame = tk.Frame(databases_frame, bg=background_color, height=80)
bawah_frame.pack(fill="x", side="bottom")

# Judul Menu
title_label = tk.Label(databases_frame, text="Menu Database", bg=background_color, fg="Black", font=("Arial", 24, "bold"))
title_label.pack(pady=10)

# Load Gambar
try:
    image_path = "assets/database.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(kiri_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(kiri_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Database di halaman Menu
database = tk.Canvas(kiri_frame, width=110, height=50, bg="white", highlightthickness=0)
database.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(database, x=5, y=5, width=100, height=40, radius=20, text="Input Data", command="")

# Load Gambar
try:
    image_path = "assets/reicon.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(tengah_tengah_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(tengah_tengah_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Riport di halaman Menu
riport = tk.Canvas(tengah_tengah_frame, width=150, height=50, bg="white", highlightthickness=0)
riport.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(riport, x=5, y=5, width=140, height=40, radius=20, text="Backup Database", command="")

# Load Gambar
try:
    image_path = "assets/abs.png"  # Pastikan jalur sesuai dengan lokasi file gambar Anda
    loaded_image = load_image(image_path, 250, 250)  # Mengatur ukuran gambar
    img_label = tk.Label(kanan_frame, image=loaded_image, bg="white")
    img_label.image = loaded_image  # Menyimpan referensi agar tidak dihapus oleh garbage collector
    img_label.pack(pady=20)
except FileNotFoundError:
    error_label = tk.Label(kanan_frame, text="Gambar tidak ditemukan!", fg="red", bg="white")
    error_label.pack(pady=20)

# Tombol Absensi Siswa di halaman Menu
absensi = tk.Canvas(kanan_frame, width=150, height=50, bg="white", highlightthickness=0)
absensi.place(relx=0.50, rely=0.85, anchor="center") 
create_rounded_button(absensi, x=5, y=5, width=140, height=40, radius=20, text="Update Presensi", command="")

# Tombol Exit di pojok kanan bawah pada bawah_frame
exited_canvas = tk.Canvas(bawah_frame, width=100, height=50, bg="white", highlightthickness=0)
exited_canvas.place(relx=0.85, rely=0.25, anchor="center") 
create_rounded_button(exited_canvas, x=5, y=5, width=90, height=40, radius=20, text="Exit", command="")

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame input data list siswa # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

base64save = None;
def loadexcel():
    global base64save
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        return  # No file selected

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Load the active sheet

        # Clear the existing table
        for row in tree.get_children():
            tree.delete(row)
        tree["columns"] = []

        # Load the header and rows
        headers = [cell.value for cell in sheet[1]]
        tree["columns"] = headers
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, width=100, anchor=tk.W)

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", "end", values=row)
            data.append(dict(zip(headers, row)))

        # Save data to a JSON file
        json_file_path = "data.json"
        with open(json_file_path, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, indent=4)
        
        # Encode JSON file content to Base64
        with open(json_file_path, "rb") as json_file:
            encoded_data = base64.b64encode(json_file.read()).decode("utf-8")
        
        # Print encoded data to the terminal
        print("Base64 Encoded Data:")
        print(encoded_data)
        base64save = encoded_data
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {e}")
def savedata():
    payload = {
        "login": tokenlogin,
        "data":"get_backup_data"
    }
    response = requests.post("http://localhost:3000/backupkelas9", data=payload)
    response = json.loads(response.text)
    if response["success"] == 1:
        df = pd.DataFrame(response["data"]["list"])
        output_file = "Backup Kelas 9 "+str(response["data"]["tahunajar"])+".xlsx"
        df.to_excel(output_file, index=False)
        payload = {
            "login": tokenlogin,
            "data":"go_insert_siswa#"+base64save
        }
        response = requests.post("http://localhost:3000/insertlistsiswa", data=payload)
        response = json.loads(response.text)
        if response["success"] == 1:
            messagebox.showinfo("Success", "Berhasil Menambahkan Data")
        else:
            if response["data"] == "err_tahunajar":
                tanyabuat = messagebox.askyesno("Konfirmasi", "Apakah ingin membuat tahun ajar baru?")
                if tanyabuat:
                    payload = {
                        "login": tokenlogin,
                        "data":"tambahtahunajar#"+base64save
                    }
                    response = requests.post("http://localhost:3000/insertlistsiswa", data=payload)
                    response = json.loads(response.text)
                    if response["success"] == 1:
                        messagebox.showinfo("Success", "Berhasil Menambahkan Data")
                    else:
                        messagebox.showwarning("Error", response["data"])
            else:
                messagebox.showwarning("Error", response["data"])
    else:
        messagebox.showwarning("Error", response["data"])

# Frame utama
input_frame = tk.Frame(root, bg="white")
input_frame.pack(fill="both", expand=True, padx=10, pady=10)

# Label Judul
title_label = tk.Label(input_frame, text="Input Data dari Excel", font=("Arial", 16, "bold"), bg="white", fg="#800080")
title_label.pack(pady=10)

# Tombol untuk memuat file Excel
load_button = ttk.Button(input_frame, text="Load Excel File", command=loadexcel)
load_button.pack(pady=10)

# Tombol untuk menyimpan data
save_button = ttk.Button(input_frame, text="Save Data", command=savedata)
save_button.pack(pady=10)

# Tabel untuk menampilkan data
tree_frame = tk.Frame(input_frame, bg="white")
tree_frame.pack(fill="both", expand=True, pady=10)

tree = ttk.Treeview(tree_frame, show="headings")
tree.pack(side="left", fill="both", expand=True)

# Scrollbar untuk tabel
scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
scrollbar.pack(side="right", fill="y")
tree.configure(yscrollcommand=scrollbar.set)
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame backup database # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame update presensi kehadiran siswa # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# Data contoh siswa
siswa_data = []
def loaddataedit():
    global siswa_data
    payload = {
        "login": tokenlogin,
        "data":"g"
    }
    response = requests.get("http://localhost:3000/listabsenhariini", data=payload)
    datalogin = json.loads(response.text)
    siswa_data = datalogin
    for siswa in siswa_data:
        display_siswa(siswa)
# Fungsi untuk mencari siswa
def search_siswa():
    search_query = search_entry.get().lower()
    for widget in result_frame.winfo_children():
        widget.destroy()

    found_siswa = [siswa for siswa in siswa_data if search_query in siswa["nama"].lower()]
    if found_siswa:
        for siswa in found_siswa:
            display_siswa(siswa)
    else:
        tk.Label(result_frame, text="Siswa tidak ditemukan.", fg="red", bg="white").pack(pady=5)
# def gantiabsen():
    
# Fungsi untuk menampilkan siswa dengan checkbox
def display_siswa(siswa):
    siswa_frame = tk.Frame(result_frame, bg="white")
    siswa_frame.pack(fill="x", pady=5)

    name_label = tk.Label(siswa_frame, text=siswa["nama"], bg="white", fg="black")
    name_label.pack(side="left", padx=10)

    # Checkbox
    masuk_var = tk.BooleanVar()
    izin_var = tk.BooleanVar()
    absen_var = tk.BooleanVar()
    if siswa['status'] == 0:
        absen_var.set(True)
    elif siswa['status'] == 1:
        masuk_var.set(True)
    elif siswa['status'] == 2:
        izin_var.set(True)
    def on_check(var):
        if var == "masuk":
            izin_var.set(False)
            absen_var.set(False)
            if masuk_var.get() == False:
                masuk_var.set(True)
            keterangan_izin.pack_forget()
        elif var == "izin":
            masuk_var.set(False)
            absen_var.set(False)
            if izin_var.get():
                keterangan_izin.pack(side="left", padx=5)
            else:
                izin_var.set(True)
        elif var == "absen":
            masuk_var.set(False)
            izin_var.set(False)
            keterangan_izin.pack_forget()
            if absen_var.get() == False:
                absen_var.set(True)
    masuk_checkbox = tk.Checkbutton(siswa_frame, text="Masuk", variable=masuk_var, command=lambda: on_check("masuk"), bg="white", fg="black")
    izin_checkbox = tk.Checkbutton(siswa_frame, text="Izin", variable=izin_var, command=lambda: on_check("izin"), bg="white",fg="black")
    absen_checkbox = tk.Checkbutton(siswa_frame, text="Absen", variable=absen_var, command=lambda: on_check("absen"), bg="white",fg="black")
    keterangan_izin = tk.Entry(siswa_frame)

    masuk_checkbox.pack(side="left", padx=5)
    izin_checkbox.pack(side="left", padx=5)
    absen_checkbox.pack(side="left", padx=5)

    # Tombol Save menggunakan ttk.Button
    def save_attendance():
        datakirim = ""
        if masuk_var.get():
            datakirim = str(siswa['nisn'])+"#1"
        elif izin_var.get():
            datakirim = str(siswa['nisn'])+"#2#"+keterangan_izin.get()
        elif absen_var.get():
            datakirim = str(siswa['nisn'])+"#0"
        payload = {
            "login": tokenlogin,
            "data":datakirim
        }
        response = requests.post("http://localhost:3000/ubahabsen", data=payload)
        print(response.text)
        datalogin = json.loads(response.text)
        messagebox.showinfo("Status Ubah Presensi", f"Nama: {siswa['nama']}\nStatus: {datalogin['data']}")

    save_button = ttk.Button(siswa_frame, text="Save", command=save_attendance, style="Purple.TButton")
    save_button.pack(side="left", padx=10)
# Styling tombol dengan ttk
style = ttk.Style()
style.configure("Purple.TButton", background="#800080", foreground="white", font=("Arial", 10))
style.map("Purple.TButton",
          background=[("active", "#5a005a")],
          foreground=[("active", "white")])

# Frame Utama
update_frame = tk.Frame(root, bg="white")
# update_frame.pack(fill="both", expand=True)

# Label Judul
title_label = tk.Label(update_frame, text="Update Kehadiran Siswa", font=("Arial", 16, "bold"), bg="white", fg="#800080")
title_label.pack(pady=10)

# Search Bar
search_frame = tk.Frame(update_frame, bg="white")
search_frame.pack(fill="x", pady=10)

search_label = tk.Label(search_frame, text="Cari Siswa:", bg="white", fg="black")
search_label.pack(side="left", padx=(0, 10))

search_entry = tk.Entry(search_frame)
search_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

search_button = ttk.Button(search_frame, text="Search", command=search_siswa, style="Purple.TButton")
search_button.pack(side="left")

# Frame untuk menampilkan hasil
result_frame = tk.Frame(update_frame, bg="white")
result_frame.pack(fill="x", pady=10)
loaddataedit()

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # Frame Riport # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

report_frame = tk.Frame(root, bg="#f8f9fa", highlightbackground="#ced4da", highlightthickness=1)
# report_frame.pack(fill="both", expand=True)

# Header label
header_label = tk.Label(report_frame, text="Export Attendance", font=("Helvetica", 18, "bold"), bg="#f8f9fa", fg="#495057")
header_label.pack(pady=(20, 10))

# Time interval label
interval_label = tk.Label(report_frame, text="Jangka Waktu", font=("Helvetica", 12, "bold"), bg="#f8f9fa", fg="#6c757d")
interval_label.pack(anchor="w", padx=20)

# Frame untuk tombol interval waktu
interval_frame = tk.Frame(report_frame, bg="#f8f9fa")
interval_frame.pack(pady=10, padx=20, anchor="w")

# Tombol interval waktu
btn_mingguan = tk.Button(interval_frame, text="Mingguan", font=("Helvetica", 10), bg="#e9ecef", fg="#333", relief="flat", borderwidth=1, padx=10, pady=5)
btn_mingguan.pack(side="left", padx=5)

btn_bulanan = tk.Button(interval_frame, text="Bulanan", font=("Helvetica", 10), bg="#e9ecef", fg="#333", relief="flat", borderwidth=1, padx=10, pady=5)
btn_bulanan.pack(side="left", padx=5)

btn_semester = tk.Button(interval_frame, text="Semester", font=("Helvetica", 10), bg="#e9ecef", fg="#333", relief="flat", borderwidth=1, padx=10, pady=5)
btn_semester.pack(side="left", padx=5)

# Frame untuk tabel data kehadiran
table_frame = tk.Frame(report_frame, bg="#f8f9fa")
table_frame.pack(pady=20, padx=20, fill="both", expand=True)

# Gaya khusus untuk tabel
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
                background="#ffffff",
                foreground="#000000",
                rowheight=25,
                fieldbackground="#ffffff")
style.map("Treeview",
          background=[("selected", "#6c757d")],
          foreground=[("selected", "white")])
style.configure("Treeview.Heading",
                font=("Helvetica", 11, "bold"),
                background="#495057",
                foreground="white",
                borderwidth=1)

# Widget Treeview untuk tabel
columns = ("No", "Nama", "Hadir", "Ijin", "Alpha")
table = ttk.Treeview(table_frame, columns=columns, show="headings", height=8)
table.pack(fill="both", expand=True)

# Tentukan kolom dan lebar tabel
for col in columns:
    table.heading(col, text=col)
    table.column(col, anchor="center", width=100)

# Warna alternatif untuk baris
table.tag_configure("evenrow", background="#f8f9fa")
table.tag_configure("oddrow", background="#ffffff")

# Masukkan data sampel ke dalam tabel
sample_data = [
    (1, "Student 1", "01/01/2023", "01/07/2023", "Hadir"),
    (2, "Student 2", "12/25/2022", "01/07/2023", "Ijin"),
    (3, "Student 3", "12/09/2022", "01/07/2023", "Alpha"),
]

for index, row in enumerate(sample_data):
    tag = "evenrow" if index % 2 == 0 else "oddrow"
    table.insert("", "end", values=row, tags=(tag,))

# Tombol Export to PDF
try:
    icon = tk.PhotoImage(file="assets/pdf_icon.png")  # Pastikan file ada
    export_btn = tk.Button(
        report_frame,
        text="Export to PDF",
        image=icon,
        compound="left",
        font=("Helvetica", 12, "bold"),
        borderwidth=1,
        bg="#e0aaff",
        fg="black",
        padx=10,
        pady=5
    )
    export_btn.pack(pady=20)
    export_btn.image = icon  # Simpan referensi agar tidak garbage-collected
except Exception as e:
    export_btn = tk.Button(
        report_frame,
        text="Export to PDF",
        font=("Helvetica", 12, "bold"),
        borderwidth=1,
        bg="#e0aaff",
        fg="black",
        padx=10,
        pady=5
    )
    export_btn.pack(pady=20)




# Menangani event ketika aplikasi ditutup
root.protocol("WM_DELETE_WINDOW", on_closing)

# Jalankan aplikasi
root.mainloop()
